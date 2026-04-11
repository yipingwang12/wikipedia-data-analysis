"""Tests for output — CSV/TSV/Excel writer."""

from __future__ import annotations

from pathlib import Path

import openpyxl

from wiki_pipeline.output import write_excel, write_results


class TestWriteResults:
    def test_csv_output(self, tmp_path: Path):
        records = [
            {"page_id": 1, "title": "Monet", "birth_date": "1840-11-14", "nationality": "French"},
            {"page_id": 2, "title": "Renoir", "birth_date": "1841-02-25", "nationality": None},
        ]
        out = write_results(records, tmp_path / "out.csv", "csv")
        content = out.read_text()
        lines = content.strip().split("\n")
        assert lines[0] == "page_id,title,birth_date,nationality"
        assert "Monet" in lines[1]
        assert lines[2].endswith(",")  # None → empty string

    def test_tsv_output(self, tmp_path: Path):
        records = [{"page_id": 1, "title": "Test"}]
        out = write_results(records, tmp_path / "out.tsv", "tsv")
        content = out.read_text()
        assert "\t" in content

    def test_empty_records(self, tmp_path: Path):
        out = write_results([], tmp_path / "empty.csv", "csv")
        assert out.read_text() == ""

    def test_none_to_empty_string(self, tmp_path: Path):
        records = [{"a": None, "b": "val"}]
        out = write_results(records, tmp_path / "out.csv", "csv")
        lines = out.read_text().strip().split("\n")
        assert lines[1] == ",val"

    def test_creates_parent_dirs(self, tmp_path: Path):
        out = write_results(
            [{"x": 1}], tmp_path / "sub" / "dir" / "out.csv", "csv"
        )
        assert out.exists()

    def test_returns_path(self, tmp_path: Path):
        out = write_results([{"x": 1}], tmp_path / "out.csv", "csv")
        assert isinstance(out, Path)
        assert out == tmp_path / "out.csv"

    def test_xlsx_via_write_results(self, tmp_path: Path):
        records = [{"page_id": 1, "title": "Monet", "birth_date": "1840"}]
        out = write_results(records, tmp_path / "out.xlsx", "xlsx")
        assert out.exists()
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        assert ws.cell(1, 1).value == "page_id"
        assert ws.cell(2, 2).value == "Monet"


class TestWriteExcel:
    def test_basic_output(self, tmp_path: Path):
        records = [
            {"page_id": 1, "title": "Monet", "birth_date": "1840"},
            {"page_id": 2, "title": "Renoir", "birth_date": "1841"},
        ]
        out = write_excel(records, tmp_path / "out.xlsx")
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        assert ws.cell(1, 1).value == "page_id"
        assert ws.cell(2, 2).value == "Monet"
        assert ws.cell(3, 2).value == "Renoir"

    def test_custom_sheet_name(self, tmp_path: Path):
        records = [{"x": 1}]
        out = write_excel(records, tmp_path / "out.xlsx", sheet_name="artists")
        wb = openpyxl.load_workbook(out)
        assert wb.active.title == "artists"

    def test_empty_records(self, tmp_path: Path):
        out = write_excel([], tmp_path / "empty.xlsx")
        assert out.exists()
        wb = openpyxl.load_workbook(out)
        assert wb.active.max_row == 1 or wb.active.max_row is None

    def test_none_values(self, tmp_path: Path):
        records = [{"a": None, "b": "val"}]
        out = write_excel(records, tmp_path / "out.xlsx")
        wb = openpyxl.load_workbook(out)
        assert wb.active.cell(2, 1).value is None
        assert wb.active.cell(2, 2).value == "val"

    def test_creates_parent_dirs(self, tmp_path: Path):
        out = write_excel([{"x": 1}], tmp_path / "sub" / "dir" / "out.xlsx")
        assert out.exists()

    def test_article_bytes_and_word_count_columns(self, tmp_path: Path):
        records = [
            {"page_id": 1, "title": "Test", "article_bytes": 5000, "word_count": 800},
        ]
        out = write_excel(records, tmp_path / "out.xlsx")
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        assert "article_bytes" in headers
        assert "word_count" in headers
        assert ws.cell(2, headers.index("article_bytes") + 1).value == 5000
        assert ws.cell(2, headers.index("word_count") + 1).value == 800
